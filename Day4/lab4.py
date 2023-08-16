from myexperiancepackage.simplefile import *
from openpyxl import *
import os

#print(read("myexperiancepackage/simplefile.py","all"))

#print(read("/etc/shadow","all"))

data=input("enter data to add : \n")
write("try1.txt",data)
print("file content is : \n",read("try1.txt","all"))

def writeXlsx(file:None,dcell,data):
    if os.path.isfile(file):
        mywb=load_workbook(file)
        mysheet=mywb.active
        mysheet[dcell]=data
        print("saving")
        mywb.save(file)

    else :
        mywb=Workbook()
        mysheet=mywb.active
        mysheet[dcell]=data
        print("saving")
        mywb.save("new.xlsx")

def readXlsx(file,cell):
    wb=load_workbook(file)
    sh=wb.active
    print(sh[cell].value)

writeXlsx(input("file path"),input("cell to add "),input("data to enter"))
readXlsx(input("enter filename "),input("enter cell to view"))

